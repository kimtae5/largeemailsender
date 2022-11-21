# Gmail Sender
# 아래 라이브러리들은 파이썬 기본 내장 라이브러리이므로 별도의 설치가 필요 없습니다.

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
from account import SMTP_USER, SMTP_PASSWORD
import time
import asyncio
from contents2 import 본문1, 본문2, 본문3, 본문4

# Google의 Gmail 같은 경우 아래의 STMP 설정을 그대로 쓰면 됩니다. 포트 번호도 바꿀 필요 없습니다.
# 이는 Google에서 설정한 것이므로 gmail, 혹은 gsuite 기준 그대로 쓰면 됩니다.
# 단, 회사 메일이나 네이버, 다음 같은 경우 다르므로 SMTP 서버에 대해 알아보는 게 좋습니다.
# 더 자세한 정보는 각 홈페이지에서 SMTP를 검색해주세요.

# Gmail 같은 경우 해당 코드가 바로는 돌아가지 않습니다. 경고창에 뜨는 google의 URL로 들어가서
# "보안 수준이 낮은 앱 허용"을 "활성화"로 바꿔주셔야 합니다.
# "보안 수준이 낮은 앱 허용"의 기본 상태는 "비활성화"입니다.
# 해당 상태를 바꾸지 않으면 코드 실행에 에러가 발생합니다.

start = time.time()
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465

# 만약 아래 메일 유효성 검사 함수에서 False가 나오면 메일을 보내지 않습니다.
async def is_valid(addr):
    import re
    if re.match('(^[a-zA-Z-0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        return True
    else:
        return False
# 이메일 보내기 함수
async def send_mail(addr, subj_layout, cont_layout, attachment=None):
    if not await is_valid(addr):
        print("Wrong email: " + addr)
        return
    
    # 텍스트 파일
    msg = MIMEMultipart("alternative")
    # 첨부파일이 있는 경우 mixed로 multipart 생성
    if attachment:
        msg = MIMEMultipart('mixed')
    msg["From"] = SMTP_USER
    msg["To"] = addr
    msg["Subject"] = subj_layout
    contents1 = cont_layout
    text = MIMEText(_text = contents1, _subtype='html', _charset = "utf-8")
    msg.attach(text)

    
    # smtp로 접속할 서버 정보를 가진 클래스변수 생성
    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    # 해당 서버로 로그인
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    # 메일 발송
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    # 닫기
    smtp.close()


# 엑셀 파일에 정리된 명단으로 한꺼번에 보낼 때
# 아래 openpyxl 라이브러리는 외부 라이브러리이므로 pip3를 통해 설치 후 사용하시기 바랍니다.
from openpyxl import load_workbook
wb = load_workbook('email2.xlsx')
ws = wb["Sheet3"]
for row in ws.iter_rows(min_row=2):
    addr = row[1].value + row[2].value
    subj_layout = ("런칭기념 이벤트 블로그 포스팅 의뢰드립니다!")
    contents = ("안녕하세요? 블로거님 :)" + 본문1 + row[3].value + 본문2 + row[5].value + 본문3 + row[4].value + 본문4)
    
    asyncio.run(send_mail(addr, subj_layout, contents))

end = time.time()
print(f'>>> 발송 처리 총 소요 시간: {end - start}')